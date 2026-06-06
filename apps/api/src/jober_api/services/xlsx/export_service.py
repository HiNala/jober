from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from sqlalchemy.ext.asyncio import AsyncSession

from jober_api.repositories.company_board import CompanyBoardRepository
from jober_api.repositories.cover_letter_angle import CoverLetterAngleRepository
from jober_api.repositories.job_target import JobTargetRepository
from jober_api.services.xlsx.normalize import export_status
from jober_api.services.xlsx.sheet_specs import (
    COMPANY_BOARDS_SPEC,
    COVER_LETTER_ANGLES_SPEC,
    JOB_LEADS_SPEC,
)


async def export_jobs_workbook(session: AsyncSession) -> bytes:
    job_repo = JobTargetRepository(session)
    board_repo = CompanyBoardRepository(session)
    angle_repo = CoverLetterAngleRepository(session)

    jobs = await job_repo.list_all(limit=10_000, offset=0)
    jobs.sort(key=lambda j: (j.rank is None, j.rank or 0))
    boards = await board_repo.list_all(limit=10_000, offset=0)
    angles = await angle_repo.list_all(limit=10_000, offset=0)

    wb = Workbook()
    default = wb.active
    wb.remove(default)

    leads_ws = wb.create_sheet(JOB_LEADS_SPEC.sheet_names[0])
    leads_headers = [col.aliases[0] for col in JOB_LEADS_SPEC.columns]
    leads_ws.append(leads_headers)
    for job in jobs:
        leads_ws.append(
            [
                job.rank,
                job.priority,
                job.company,
                job.role,
                job.fit_lane,
                job.stage_signal,
                job.location_work_style,
                job.why_fit,
                job.cover_letter_hook,
                job.public_contact,
                job.direct_apply_url,
                job.company_careers_url,
                job.source_note,
                job.verified_date,
                export_status(job.status),
                job.applied_date,
                job.follow_up_date,
                job.notes,
            ]
        )

    boards_ws = wb.create_sheet(COMPANY_BOARDS_SPEC.sheet_names[0])
    boards_headers = [col.aliases[0] for col in COMPANY_BOARDS_SPEC.columns]
    boards_ws.append(boards_headers)
    for board in boards:
        boards_ws.append(
            [
                board.priority,
                board.company_board,
                board.representative_roles,
                board.stage_signal,
                board.why_save,
                board.company_careers_url,
                board.last_checked,
                board.notes,
            ]
        )

    angles_ws = wb.create_sheet(COVER_LETTER_ANGLES_SPEC.sheet_names[0])
    angles_headers = [col.aliases[0] for col in COVER_LETTER_ANGLES_SPEC.columns]
    angles_ws.append(angles_headers)
    for angle in angles:
        angles_ws.append([angle.use_case, angle.template])

    summary_ws = wb.create_sheet("Summary")
    summary_ws.append(["Jober export", "Round-trip status fields reflect app state"])
    refresh_ws = wb.create_sheet("Refresh Sources")
    refresh_ws.append(["Source", "Exported from Jober database"])

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
