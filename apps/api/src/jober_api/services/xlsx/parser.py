from __future__ import annotations

import re
from dataclasses import dataclass
from io import BytesIO

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from jober_api.services.xlsx.column_mapping import ResolvedColumnMapping, resolve_sheet_mapping
from jober_api.services.xlsx.sheet_specs import (
    COMPANY_BOARDS_SPEC,
    COVER_LETTER_ANGLES_SPEC,
    JOB_LEADS_SPEC,
    SheetSpec,
)


def _normalize_sheet_name(name: str) -> str:
    return re.sub(r"\s+", " ", name.strip().lower())


def find_sheet_name(workbook: Workbook, spec: SheetSpec) -> str | None:
    targets = {_normalize_sheet_name(n) for n in spec.sheet_names}
    for name in workbook.sheetnames:
        if _normalize_sheet_name(name) in targets:
            return str(name)
    for name in workbook.sheetnames:
        norm = _normalize_sheet_name(name)
        for target in targets:
            if target in norm or norm in target:
                return str(name)
    return None


def _sheet_rows(ws: Worksheet, max_rows: int = 5000) -> list[tuple[object, ...]]:
    rows: list[tuple[object, ...]] = []
    for idx, row in enumerate(ws.iter_rows(values_only=True)):
        if idx >= max_rows:
            break
        rows.append(tuple(row))
    return rows


@dataclass(frozen=True)
class ParsedSheet:
    sheet_name: str
    mapping: ResolvedColumnMapping
    records: list[dict[str, object]]


@dataclass(frozen=True)
class WorkbookParseResult:
    job_leads: ParsedSheet | None
    company_boards: ParsedSheet | None
    cover_letter_angles: ParsedSheet | None
    metadata: dict[str, list[list[object]]]


def _records_from_sheet(
    rows: list[tuple[object, ...]],
    mapping: ResolvedColumnMapping,
    spec: SheetSpec,
) -> list[dict[str, object]]:
    header_row = rows[mapping.header_row]
    header_index = {
        str(cell).strip(): idx for idx, cell in enumerate(header_row) if cell is not None
    }

    records: list[dict[str, object]] = []
    for row_num, row in enumerate(rows[mapping.header_row + 1 :], start=mapping.header_row + 2):
        record: dict[str, object] = {"_row": row_num}
        empty = True
        for col in spec.columns:
            header = mapping.fields.get(col.field)
            if not header:
                record[col.field] = None
                continue
            idx = header_index.get(header)
            value = row[idx] if idx is not None and idx < len(row) else None
            if value not in (None, ""):
                empty = False
            record[col.field] = value
        if empty:
            continue
        records.append(record)
    return records


def _read_metadata_sheet(ws: Worksheet, max_rows: int = 30) -> list[list[object]]:
    meta: list[list[object]] = []
    for idx, row in enumerate(ws.iter_rows(values_only=True)):
        if idx >= max_rows:
            break
        if any(cell not in (None, "") for cell in row):
            meta.append(list(row))
    return meta


def parse_workbook_bytes(data: bytes) -> WorkbookParseResult:
    wb = load_workbook(BytesIO(data), read_only=True, data_only=True)

    def parse_entity(spec: SheetSpec) -> ParsedSheet | None:
        sheet_name = find_sheet_name(wb, spec)
        if not sheet_name:
            return None
        ws = wb[sheet_name]
        rows = _sheet_rows(ws)
        mapping = resolve_sheet_mapping(sheet_name, rows, spec)
        if mapping is None:
            return None
        records = _records_from_sheet(rows, mapping, spec)
        return ParsedSheet(sheet_name=sheet_name, mapping=mapping, records=records)

    metadata: dict[str, list[list[object]]] = {}
    for name in wb.sheetnames:
        norm = _normalize_sheet_name(name)
        if norm in {"summary", "refresh sources"}:
            metadata[name] = _read_metadata_sheet(wb[name])

    return WorkbookParseResult(
        job_leads=parse_entity(JOB_LEADS_SPEC),
        company_boards=parse_entity(COMPANY_BOARDS_SPEC),
        cover_letter_angles=parse_entity(COVER_LETTER_ANGLES_SPEC),
        metadata=metadata,
    )
