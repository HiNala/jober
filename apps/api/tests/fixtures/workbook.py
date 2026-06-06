from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook

from jober_api.services.xlsx.sheet_specs import (
    COMPANY_BOARDS_SPEC,
    COVER_LETTER_ANGLES_SPEC,
    JOB_LEADS_SPEC,
)


def build_sample_workbook(
    *,
    job_count: int = 3,
    board_count: int = 2,
    angle_count: int = 2,
) -> bytes:
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    leads = wb.create_sheet(JOB_LEADS_SPEC.sheet_names[0])
    leads.append([col.aliases[0] for col in JOB_LEADS_SPEC.columns])
    for i in range(1, job_count + 1):
        leads.append(
            [
                i,
                "A",
                f"Company {i}",
                f"Role {i}",
                "IC",
                "Series B",
                "Remote",
                "Fit reason",
                "Hook",
                f"hiring{i}@example.com",
                f"https://jobs.lever.co/acme/role-{i}",
                f"https://boards.greenhouse.io/acme/jobs/{i}",
                "Verified on web",
                "2025-01-01",
                "Not started",
                None,
                None,
                f"Note {i}",
            ]
        )

    boards = wb.create_sheet(COMPANY_BOARDS_SPEC.sheet_names[0])
    boards.append([col.aliases[0] for col in COMPANY_BOARDS_SPEC.columns])
    for i in range(1, board_count + 1):
        boards.append(
            [
                "B",
                f"Board Co {i}",
                f"Engineer {i}",
                "Growth",
                "Strong culture",
                f"https://jobs.ashbyhq.com/board{i}",
                "2025-02-01",
                f"Board note {i}",
            ]
        )

    angles = wb.create_sheet(COVER_LETTER_ANGLES_SPEC.sheet_names[0])
    angles.append([col.aliases[0] for col in COVER_LETTER_ANGLES_SPEC.columns])
    for i in range(1, angle_count + 1):
        angles.append([f"Use case {i}", f"Template body {i}"])

    summary = wb.create_sheet("Summary")
    summary.append(["Tracker summary", "Informational only"])
    summary.append(["Total leads", job_count])

    refresh = wb.create_sheet("Refresh Sources")
    refresh.append(["Source", "Frequency"])

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
