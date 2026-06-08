from __future__ import annotations

import os
from io import BytesIO

import pytest
from openpyxl import load_workbook
from sqlalchemy import select

from jober_api.auth.constants import DEFAULT_DEV_TENANT_ID
from jober_api.models.enums import JobTargetStatus
from jober_api.models.job_target import JobTarget
from jober_api.repositories.job_target import JobTargetRepository
from jober_api.services.xlsx.export_service import export_jobs_workbook
from jober_api.services.xlsx.import_service import import_jobs_workbook
from tests.fixtures.workbook import build_sample_workbook

pytestmark = pytest.mark.skipif(
    os.getenv("CI") != "true" and os.getenv("RUN_DB_TESTS") != "1",
    reason="requires Postgres",
)


def _status_for_company(exported: bytes, company: str) -> str | None:
    wb = load_workbook(BytesIO(exported), read_only=True, data_only=True)
    ws = wb["Direct Job Leads"]
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    status_idx = headers.index("Status")
    company_idx = headers.index("Company")
    for row in rows[1:]:
        if row[company_idx] == company:
            return str(row[status_idx]) if row[status_idx] is not None else None
    return None


@pytest.mark.asyncio
async def test_export_reflects_app_status_after_update(db_session, truncate_tables) -> None:
    data = build_sample_workbook(job_count=2, board_count=1, angle_count=1)
    await import_jobs_workbook(db_session, data, tenant_id=DEFAULT_DEV_TENANT_ID, dry_run=False)

    target = (
        await db_session.execute(select(JobTarget).where(JobTarget.company == "Company 1"))
    ).scalar_one()
    target.status = JobTargetStatus.APPLIED
    target.notes = "Submitted via Jober"
    await db_session.commit()

    exported = await export_jobs_workbook(db_session, DEFAULT_DEV_TENANT_ID)
    assert _status_for_company(exported, "Company 1") == "Applied"
    assert _status_for_company(exported, "Company 2") == "Not started"


@pytest.mark.asyncio
async def test_reimport_preserves_app_status_when_runs_exist(db_session, truncate_tables) -> None:
    from jober_api.models.application_run import ApplicationRun
    from jober_api.models.enums import RunPolicy, RunStatus

    data = build_sample_workbook(job_count=1, board_count=0, angle_count=0)
    await import_jobs_workbook(db_session, data, tenant_id=DEFAULT_DEV_TENANT_ID, dry_run=False)

    target = (
        await db_session.execute(select(JobTarget).where(JobTarget.company == "Company 1"))
    ).scalar_one()
    target.status = JobTargetStatus.IN_PROGRESS
    db_session.add(
        ApplicationRun(
            tenant_id=target.tenant_id,
            job_target_id=target.id,
            status=RunStatus.QUEUED,
            policy=RunPolicy.REVIEW_BEFORE_SUBMIT,
        )
    )
    await db_session.commit()

    # Workbook still says "Not started" for Company 1
    report = await import_jobs_workbook(db_session, data, tenant_id=DEFAULT_DEV_TENANT_ID, dry_run=False)
    assert report.job_targets.updated == 1
    repo = JobTargetRepository(db_session)
    refreshed = await repo.get(target.id)
    assert refreshed is not None
    assert refreshed.status == JobTargetStatus.IN_PROGRESS
    assert any(w.code == "status_preserved" for w in report.warnings)
